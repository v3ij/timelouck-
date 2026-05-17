import openpyxl

def xlsx_to_txt(xlsx_path, txt_path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    fullText = []
    
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        fullText.append(f"=== Sheet: {sheet_name} ===")
        for row in sheet.iter_rows(values_only=True):
            # Convert row to string
            row_str = []
            for cell in row:
                if cell is None:
                    row_str.append("")
                else:
                    row_str.append(str(cell).strip())
            # Skip completely empty rows
            if any(row_str):
                fullText.append(" | ".join(row_str))
        fullText.append("\n")
        
    with open(txt_path, 'w', encoding='utf-8') as f:
        f.write('\n'.join(fullText))

xlsx_to_txt(
    "c:\\Users\\ayoub\\OneDrive\\سطح المكتب\\africano project\\Time-Wallet_System_Modules_Requirementd_Gantt_Chart.xlsx",
    "c:\\Users\\ayoub\\OneDrive\\سطح المكتب\\africano project\\scratch\\Time-Wallet_System_Modules_Requirementd_Gantt_Chart.txt"
)
print("Xlsx conversion done!")
